from __future__ import annotations

import json
import logging
import os
import tempfile
from datetime import timedelta
from decimal import Decimal
from io import BytesIO
from threading import Thread

from dateutil.relativedelta import relativedelta
from flask import current_app, has_request_context, request
from flask_mail import Message
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .extensions import db, mail
from .models import (
    AuditLog,
    AuctionBid,
    ChitCycle,
    ChitGroup,
    GroupMembership,
    InstallmentSchedule,
    LedgerEntry,
    Member,
    Payment,
    today_ist,
)


def log_audit(actor_id: int | None, action: str, entity_type: str, entity_id: str | int, details: dict | None = None) -> None:
    details_payload = json.dumps(details or {}, default=str)
    audit_log = AuditLog(
        actor_user_id=actor_id,
        action=action,
        entity_type=entity_type,
        entity_id=str(entity_id),
        details=details_payload,
        ip_address=request.remote_addr if has_request_context() else None,
    )
    db.session.add(audit_log)


def generate_installment_schedule(group: ChitGroup, actor_id: int | None = None) -> None:
    group.schedules.clear()
    for round_number in range(1, group.total_members + 1):
        due_date = group.start_date + relativedelta(months=round_number - 1)
        schedule = InstallmentSchedule(
            round_number=round_number,
            due_date=due_date,
            expected_amount=group.monthly_amount,
            created_by=actor_id,
            updated_by=actor_id,
        )
        group.schedules.append(schedule)


def generate_group_cycles(group: ChitGroup, actor_id: int | None = None) -> None:
    group.cycles.clear()
    for cycle_number in range(1, group.total_members + 1):
        due_date = group.start_date + relativedelta(months=cycle_number - 1)
        auction_date = due_date + timedelta(days=max(group.auction_day - 1, 0))
        cycle = ChitCycle(
            cycle_number=cycle_number,
            due_date=due_date,
            auction_date=auction_date,
            status="Open" if cycle_number == group.current_round else "Scheduled",
            expected_collection=Decimal(str(group.pool_value)),
            created_by=actor_id,
            updated_by=actor_id,
        )
        group.cycles.append(cycle)


def enroll_member_in_group(
    member: Member,
    group: ChitGroup,
    actor_id: int | None = None,
    *,
    is_primary: bool = False,
    member_number: str | None = None,
    share_units: float = 1.0,
    slot_number: int | None = None,
) -> GroupMembership:
    if slot_number is None:
        existing_slots = [
            membership.slot_number
            for membership in member.memberships
            if membership.group_id == group.id and not membership.deleted
        ]
        slot_number = max(existing_slots, default=0) + 1

    membership = GroupMembership(
        member=member,
        group=group,
        is_primary=is_primary,
        member_number=member_number,
        share_units=share_units,
        slot_number=slot_number,
        created_by=actor_id,
        updated_by=actor_id,
    )
    if is_primary:
        member.group = group
    db.session.add(membership)
    log_audit(
        actor_id,
        "membership.created",
        "GroupMembership",
        "new",
        {"member": member.name, "group": group.name, "slot_number": slot_number, "share_units": share_units},
    )
    return membership


def compute_penalty(outstanding_amount: float, due_date) -> float:
    if not due_date or due_date >= today_ist() or outstanding_amount <= 0:
        return 0.0
    rate = float(current_app.config.get("PENALTY_RATE_PERCENT", 2.0))
    months_late = max((today_ist().year - due_date.year) * 12 + today_ist().month - due_date.month, 1)
    return round(outstanding_amount * (rate / 100) * months_late, 2)


def create_payment(member: Member, membership: GroupMembership, amount: float, actor_id: int | None = None) -> Payment:
    cycle = membership.current_cycle
    due_date = cycle.due_date if cycle else None
    expected_amount = membership.expected_amount
    penalty_amount = compute_penalty(max(expected_amount - membership.current_cycle_paid_amount, 0), due_date)
    status = "Paid" if amount >= expected_amount else "Partial"
    payment = Payment(
        member=member,
        group=membership.group,
        membership=membership,
        cycle=cycle,
        amount=amount,
        expected_amount=expected_amount,
        penalty_amount=penalty_amount,
        status=status,
        due_date=due_date,
        created_by=actor_id,
        updated_by=actor_id,
    )
    member.paid_amount = round(float(member.paid_amount) + amount, 2)
    membership.penalty_balance = round(float(membership.penalty_balance) + penalty_amount, 2)
    if cycle:
        cycle.collected_amount = round(float(cycle.collected_amount) + amount, 2)
        cycle.penalty_total = round(float(cycle.penalty_total) + penalty_amount, 2)

    db.session.add(payment)
    db.session.flush()

    db.session.add(
        LedgerEntry(
            group=membership.group,
            member=member,
            membership=membership,
            cycle=cycle,
            payment=payment,
            entry_type="Payment",
            amount=amount,
            description=f"Collection received for {membership.group.name}",
            created_by=actor_id,
            updated_by=actor_id,
        )
    )
    if penalty_amount:
        db.session.add(
            LedgerEntry(
                group=membership.group,
                member=member,
                membership=membership,
                cycle=cycle,
                payment=payment,
                entry_type="Penalty",
                amount=penalty_amount,
                description=f"Late penalty applied for {membership.group.name}",
                created_by=actor_id,
                updated_by=actor_id,
            )
        )

    log_audit(actor_id, "payment.created", "Payment", payment.id, {"amount": amount, "member": member.name})
    return payment


def close_auction(cycle: ChitCycle, actor_id: int | None = None) -> AuctionBid | None:
    open_bids = [bid for bid in cycle.bids if not bid.deleted]
    if not open_bids:
        return None

    winning_bid = min(open_bids, key=lambda bid: float(bid.bid_amount))
    _apply_cycle_winner(cycle, winning_bid, float(winning_bid.bid_amount), actor_id)
    log_audit(actor_id, "auction.closed", "ChitCycle", cycle.id, {"winner_bid": float(winning_bid.bid_amount)})
    return winning_bid


def assign_cycle_winner(
    cycle: ChitCycle,
    membership: GroupMembership,
    payout_amount: float,
    actor_id: int | None = None,
    *,
    note: str | None = None,
) -> AuctionBid:
    winning_bid = AuctionBid(
        cycle=cycle,
        membership=membership,
        bid_amount=payout_amount,
        note=note,
        is_winner=True,
        created_by=actor_id,
        updated_by=actor_id,
    )
    db.session.add(winning_bid)
    db.session.flush()
    _apply_cycle_winner(cycle, winning_bid, payout_amount, actor_id)
    log_audit(
        actor_id,
        "auction.winner_selected",
        "ChitCycle",
        cycle.id,
        {"member": membership.member.name, "slot_number": membership.slot_number, "payout_amount": payout_amount},
    )
    return winning_bid


def _apply_cycle_winner(cycle: ChitCycle, winning_bid: AuctionBid, payout_amount: float, actor_id: int | None = None) -> None:
    winner_eligible_prize = winning_bid.membership.eligible_prize_amount
    winning_amount = min(float(payout_amount), winner_eligible_prize)
    discount_amount = max(winner_eligible_prize - winning_amount, 0.0)
    active_memberships = [membership for membership in cycle.group.memberships if membership.status == "Active" and not membership.deleted]
    total_active_shares = sum(float(membership.share_units) for membership in active_memberships) or 1.0

    for bid in cycle.bids:
        if bid.deleted:
            continue
        bid.is_winner = bid.id == winning_bid.id
        bid.discount_amount = discount_amount if bid.id == winning_bid.id else 0

    cycle.status = "Closed"
    cycle.winner_membership = winning_bid.membership
    cycle.winning_bid_amount = winning_amount
    cycle.discount_amount = discount_amount
    cycle.dividend_per_member = round(discount_amount / total_active_shares, 2)

    for membership in active_memberships:
        member_dividend = round(float(cycle.dividend_per_member) * float(membership.share_units), 2)
        membership.total_dividend = round(float(membership.total_dividend) + member_dividend, 2)
        db.session.add(
            LedgerEntry(
                group=cycle.group,
                member=membership.member,
                membership=membership,
                cycle=cycle,
                auction_bid=winning_bid if membership.id == winning_bid.membership_id else None,
                entry_type="Dividend",
                amount=member_dividend,
                description=f"Cycle {cycle.cycle_number} discount distributed",
                created_by=actor_id,
                updated_by=actor_id,
            )
        )

    db.session.add(
        LedgerEntry(
            group=cycle.group,
            member=winning_bid.membership.member,
            membership=winning_bid.membership,
            cycle=cycle,
            auction_bid=winning_bid,
            entry_type="AuctionPayout",
            amount=winning_amount,
            description=f"Auction payout for cycle {cycle.cycle_number}",
            created_by=actor_id,
            updated_by=actor_id,
            )
        )

def pending_memberships(group_id: int | None = None) -> list[GroupMembership]:
    query = GroupMembership.query.filter_by(status="Active", deleted=False)
    if group_id:
        query = query.filter_by(group_id=group_id)
    return [membership for membership in query.order_by(GroupMembership.id.asc()).all() if membership.payment_status != "Paid"]


def build_dashboard_metrics():
    groups = ChitGroup.query.filter_by(deleted=False).all()
    payments = Payment.query.filter_by(deleted=False).all()
    memberships = GroupMembership.query.filter_by(deleted=False).all()
    cycles = ChitCycle.query.filter_by(deleted=False).all()
    total_collections = round(sum(float(payment.amount) for payment in payments), 2)
    total_penalties = round(sum(float(payment.penalty_amount) for payment in payments), 2)
    overdue_memberships = [membership for membership in memberships if membership.payment_status == "Overdue"]
    pending_membership_count = sum(1 for membership in memberships if membership.payment_status == "Pending")
    active_groups = sum(1 for group in groups if group.active_membership_count)
    closed_auctions = sum(1 for cycle in cycles if cycle.status == "Closed")

    return {
        "total_collections": total_collections,
        "total_penalties": total_penalties,
        "pending_payments": pending_membership_count,
        "overdue_memberships": overdue_memberships,
        "profit": round(total_penalties, 2),
        "active_groups": active_groups,
        "closed_auctions": closed_auctions,
    }


def queue_payment_notifications(member: Member, payment: Payment) -> None:
    app = current_app._get_current_object()
    payload = {
        "member_name": member.name,
        "member_email": member.email,
        "member_phone": member.phone,
        "amount": float(payment.amount),
        "timestamp": payment.formatted_timestamp,
    }
    Thread(target=_send_notifications, args=(app, payload), daemon=True).start()


def _send_notifications(app, payload: dict) -> None:
    with app.app_context():
        _send_sms(app, payload)
        _send_email(payload)


def _send_sms(app, payload: dict) -> None:
    if not (
        payload["member_phone"]
        and app.config.get("TWILIO_ACCOUNT_SID")
        and app.config.get("TWILIO_AUTH_TOKEN")
        and app.config.get("TWILIO_PHONE_NUMBER")
    ):
        return
    try:
        from twilio.rest import Client

        client = Client(app.config["TWILIO_ACCOUNT_SID"], app.config["TWILIO_AUTH_TOKEN"])
        client.messages.create(
            body=f"Payment received from {payload['member_name']}: Rs {payload['amount']:.2f} on {payload['timestamp']}.",
            from_=app.config["TWILIO_PHONE_NUMBER"],
            to=payload["member_phone"],
        )
    except Exception:
        app.logger.exception("SMS notification failed")


def _send_email(payload: dict) -> None:
    if not payload["member_email"]:
        return
    try:
        message = Message(
            subject="Payment received",
            recipients=[payload["member_email"]],
            body=(
                f"Hello {payload['member_name']},\n\n"
                f"We received your payment of Rs {payload['amount']:.2f} on {payload['timestamp']}.\n"
                "Thank you."
            ),
        )
        mail.send(message)
    except Exception:
        logging.getLogger(__name__).exception("Email notification failed")


def build_payment_excel(payments: list[Payment]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Payments"
    headers = ["Payment ID", "Member", "Group", "Cycle", "Status", "Amount", "Penalty", "Timestamp"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for payment in payments:
        sheet.append(
            [
                payment.id,
                payment.member.name,
                payment.group.name if payment.group else "Unassigned",
                payment.cycle.cycle_number if payment.cycle else "-",
                payment.status,
                float(payment.amount),
                float(payment.penalty_amount),
                payment.formatted_timestamp,
            ]
        )

    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[column[0].column_letter].width = max_length

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_receipt_pdf(payment: Payment) -> BytesIO:
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    temp_file.close()

    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 18)
        pdf.cell(0, 12, "CHIT FUND RECEIPT", ln=1, align="C")
        pdf.ln(6)
        pdf.set_font("Helvetica", size=12)
        pdf.cell(0, 10, f"Member: {payment.member.name}", ln=1)
        pdf.cell(0, 10, f"Group: {payment.group.name if payment.group else 'Unassigned'}", ln=1)
        pdf.cell(0, 10, f"Share Type: {payment.membership.share_label if payment.membership else 'Full Chit'}", ln=1)
        pdf.cell(0, 10, f"Cycle: {payment.cycle.cycle_number if payment.cycle else '-'}", ln=1)
        pdf.cell(0, 10, f"Amount Paid: Rs {float(payment.amount):.2f}", ln=1)
        pdf.cell(0, 10, f"Expected Monthly Due: Rs {float(payment.expected_amount):.2f}", ln=1)
        pdf.cell(0, 10, f"Penalty: Rs {float(payment.penalty_amount):.2f}", ln=1)
        pdf.cell(0, 10, f"Status: {payment.status}", ln=1)
        pdf.cell(0, 10, f"Date: {payment.formatted_timestamp}", ln=1)
        if payment.cycle and payment.cycle.winner_membership:
            winner = payment.cycle.winner_membership
            pdf.ln(4)
            pdf.set_font("Helvetica", "B", 13)
            pdf.cell(0, 10, "Auction Winner Details", ln=1)
            pdf.set_font("Helvetica", size=12)
            pdf.cell(0, 10, f"Winner: {winner.member.name}", ln=1)
            pdf.cell(0, 10, f"Winner Group: {winner.group.name}", ln=1)
            pdf.cell(0, 10, f"Winner Share: {winner.share_label}", ln=1)
            pdf.cell(0, 10, f"Eligible Prize Amount: Rs {winner.eligible_prize_amount:.2f}", ln=1)
            pdf.cell(0, 10, f"Winning Bid: Rs {float(payment.cycle.winning_bid_amount or 0):.2f}", ln=1)
            pdf.cell(0, 10, f"Final Payout: Rs {float(payment.cycle.winner_payout_amount or 0):.2f}", ln=1)
            pdf.cell(0, 10, f"Discount: Rs {float(payment.cycle.discount_amount or 0):.2f}", ln=1)
            pdf.cell(0, 10, f"Dividend Per Member: Rs {float(payment.cycle.dividend_per_member or 0):.2f}", ln=1)
        pdf.cell(0, 10, "Thank you for your payment.", ln=1)
        pdf.output(temp_file.name)

        with open(temp_file.name, "rb") as pdf_handle:
            output = BytesIO(pdf_handle.read())
        output.seek(0)
        return output
    finally:
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)


def build_member_history_pdf(member: Member) -> BytesIO:
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    temp_file.close()

    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 18)
        pdf.cell(0, 12, "MEMBER PAYMENT HISTORY", ln=1, align="C")
        pdf.ln(4)
        pdf.set_font("Helvetica", size=12)
        pdf.cell(0, 8, f"Member: {member.name}", ln=1)
        pdf.cell(0, 8, f"Phone: {member.phone or '-'}", ln=1)
        pdf.cell(0, 8, f"Email: {member.email or '-'}", ln=1)
        pdf.cell(0, 8, f"Total Paid: Rs {float(member.paid_amount):.2f}", ln=1)
        pdf.cell(0, 8, f"Total Due: Rs {member.due_amount:.2f}", ln=1)
        pdf.ln(4)

        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 10, "Membership Summary", ln=1)
        pdf.set_font("Helvetica", size=11)
        if member.memberships:
            for membership in member.memberships:
                if membership.deleted:
                    continue
                pdf.multi_cell(
                    0,
                    8,
                    (
                        f"Group: {membership.group.name} | Share: {membership.share_label} | "
                        f"Monthly Due: Rs {membership.expected_amount:.2f} | "
                        f"Eligible Prize: Rs {membership.eligible_prize_amount:.2f} | "
                        f"Outstanding: Rs {membership.outstanding_amount:.2f}"
                    ),
                )
        else:
            pdf.cell(0, 8, "No memberships found.", ln=1)

        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 10, "Winning History", ln=1)
        pdf.set_font("Helvetica", size=11)
        winning_cycles = [
            cycle for membership in member.memberships for cycle in membership.cycles_won if not cycle.deleted and cycle.status == "Closed"
        ]
        if winning_cycles:
            for cycle in sorted(winning_cycles, key=lambda item: item.cycle_number):
                winner = cycle.winner_membership
                eligible_prize = winner.eligible_prize_amount if winner else 0.0
                share_label = winner.share_label if winner else "-"
                pdf.multi_cell(
                    0,
                    8,
                    (
                        f"Won Month/Cycle: {cycle.cycle_number} | Group: {cycle.group.name} | "
                        f"Share: {share_label} | "
                        f"Eligible Prize: Rs {eligible_prize:.2f} | "
                        f"Winning Bid: Rs {float(cycle.winning_bid_amount or 0):.2f} | "
                        f"Final Payout: Rs {float(cycle.winner_payout_amount or 0):.2f} | "
                        f"Discount: Rs {float(cycle.discount_amount or 0):.2f}"
                    ),
                )
        else:
            pdf.cell(0, 8, "Winning History: Not won yet", ln=1)

        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 10, "Payment History", ln=1)
        pdf.set_font("Helvetica", size=10)
        pdf.cell(40, 8, "Group", 1)
        pdf.cell(18, 8, "Cycle", 1)
        pdf.cell(28, 8, "Share", 1)
        pdf.cell(28, 8, "Status", 1)
        pdf.cell(30, 8, "Amount", 1)
        pdf.cell(42, 8, "Date", 1, ln=1)
        for payment in sorted(member.payments, key=lambda item: item.timestamp or item.created_at, reverse=True):
            pdf.cell(40, 8, (payment.group.name if payment.group else "-")[:18], 1)
            pdf.cell(18, 8, str(payment.cycle.cycle_number if payment.cycle else "-"), 1)
            pdf.cell(28, 8, (payment.membership.share_label if payment.membership else "Full")[:12], 1)
            pdf.cell(28, 8, payment.status[:12], 1)
            pdf.cell(30, 8, f"Rs {float(payment.amount):.0f}", 1)
            pdf.cell(42, 8, payment.formatted_timestamp[:19], 1, ln=1)
        if not member.payments:
            pdf.cell(0, 8, "No payments found.", ln=1)

        pdf.output(temp_file.name)
        with open(temp_file.name, "rb") as pdf_handle:
            output = BytesIO(pdf_handle.read())
        output.seek(0)
        return output
    finally:
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
